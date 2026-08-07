"""Excel parser — each sheet becomes a separate section."""

import openpyxl
from typing import Any, Dict

from .base import BaseParser


class ExcelParser(BaseParser):

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]

    async def parse(self, file_path: str) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sections = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(c.strip() for c in cells):  # Skip empty rows
                    rows.append(" | ".join(cells))

            if rows:
                header_sep = "---|" * len(rows[0].split("|")) if rows else ""
                table_text = f"<table>\n## Sheet: {sheet_name}\n{rows[0]}\n{header_sep}\n"
                table_text += "\n".join(rows[1:]) + "\n</table>"
                sections.append(table_text)

        wb.close()

        return {
            "text": "\n\n".join(sections),
            "metadata": {"sheets": wb.sheetnames},
            "tables": [],
        }
